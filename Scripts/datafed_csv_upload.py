"""
upload_to_datafed.py

Uploads preprocessed data and metadata to DataFed.

Usage:
    python upload_to_datafed.py --config config/DataFedSettings.cfg --schema schema.json [--create-only] [--reads-dir Data/reads]

Notes:
    - Dry-run via --dry-run flag or ECOKMER_DRY_RUN=1 env var (skips auth & remote ops).
    - Schema drives column mapping, title field, dependency relationships, and optional update mode.
"""

import csv
import json
import os
import argparse
import logging
import getpass
import configparser

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None
try:
    from datafed.CommandLib import API  # type: ignore
except Exception:  # pragma: no cover
    API = None

CONFIG_SECTION = "DataFed Project Settings"
DRY_RUN_ENV = "ECOKMER_DRY_RUN"  # GOTCHA: Environment flag only; CLI --dry-run sets an attribute flag on is_dry_run.


def setup_logging(log_file):
    """Configure file + console logging (idempotent for repeated CLI invocations)."""
    logging.basicConfig(
        filename=log_file,
        filemode="a",
        format="%(asctime)s - %(levelname)s - %(message)s",
        level=logging.INFO,
    )
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter("%(levelname)s - %(message)s")
    console.setFormatter(formatter)
    logging.getLogger("").addHandler(console)


def create_config(filename):
    """Interactively create an initial config file with endpoint, batch file, context, parent collection."""
    endpoint_uuid = input("Enter your Globus Endpoint UUID: ").strip()
    batch_file = input(
        "Enter your batch CSV file path (e.g., data/inputs/batch_file.csv): "
    ).strip()
    context = input("Enter your DataFed context (e.g., p/brave): ").strip()
    parent_collection = input(
        "Enter your DataFed parent collection ID (e.g., root, thrust3, 'c/511488914'): "
    ).strip()
    if not os.path.exists(batch_file):
        print(f"Batch file does not exist: {batch_file}")
        raise RuntimeError(f"Batch file does not exist: {batch_file}")
    config = configparser.ConfigParser()
    config["DataFed Project Settings"] = {
        "globus_uuid": endpoint_uuid,
        "batch_file": batch_file,
        "context": context,
        "parent_collection": parent_collection,
    }
    with open(filename, "w") as f:
        config.write(f)
    print(f"Configuration written to {filename}")


def read_config(filename):
    """Read ini-style config file and return ConfigParser object."""
    config = configparser.ConfigParser()
    config.read(filename)
    return config


def load_schema(schema_file):
    """Load JSON schema describing mapping, title source, dependencies, update mode, and optional sheet name."""
    if not os.path.exists(schema_file):
        logging.error(f"Schema file not found: {schema_file}")
        raise FileNotFoundError(f"Schema file not found: {schema_file}")
    with open(schema_file, "r") as f:
        schema = json.load(f)
    return schema


def transform_title(raw_title, transform_rule):
    """Apply simple title transform rules (currently only 'remove_extension')."""
    if transform_rule == "remove_extension":
        return raw_title.partition(".")[0]
    return raw_title


def parse_data_with_schema(batch_file, schema, read_row_start=1):
    """
    Parses the input file (CSV or XLSX) based on the schema.
    For XLSX files, if sheet_name is provided, that sheet is used; otherwise, the first sheet is read.
    Returns a list of records (dictionaries) using the schema to map columns.
    """
    rows = []
    ext = os.path.splitext(batch_file)[1].lower()
    if ext == ".xlsx":
        wb = load_workbook(batch_file, data_only=True)
        sheet_name = schema["sheet_name"]
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Using {sheet_name} from spreadsheet")
        else:
            ws = wb.active  # Use first sheet if sheet_name not provided
        # Define where the header row is, default 1
        header_row = next(ws.iter_rows(min_row=read_row_start, max_row=read_row_start))
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in header_row
        ]

        # Process the remaining rows
        for row in ws.iter_rows(min_row=read_row_start + 1, values_only=True):
            row_dict = {}
            for i in range(len(headers)):
                # Convert value to string and strip if not None
                cell_value = row[i]
                row_dict[headers[i]] = (
                    str(cell_value).strip() if cell_value is not None else ""
                )
            # Build record based on schema mapping
            record = {}
            for field, props in schema["properties"].items():
                colname = props.get("csv_col", field)
                record[field] = row_dict.get(colname, "")
            rows.append(record)

    elif ext == ".csv":
        with open(batch_file, "r", newline="") as csvfile:
            reader = csv.DictReader(csvfile)
            reader.fieldnames = [header.strip() for header in reader.fieldnames]
            for row in reader:
                record = {}
                for field, props in schema["properties"].items():
                    colname = props.get("csv_col", field)
                    record[field] = row.get(colname, "").strip()
                rows.append(record)
    elif ext == ".txt":
        # We'll read the first line as headers, then subsequent lines as data
        with open(batch_file, "r") as f:
            lines = [line.strip() for line in f if line.strip()]
            if not lines:
                return rows
            # Assume first line is header
            headers = lines[0].split()
            for data_line in lines[1:]:
                columns = data_line.split()
                row_dict = {}
                for i, h in enumerate(headers):
                    row_dict[h] = columns[i] if i < len(columns) else ""
                # Build record based on schema mapping
                record = {}
                for field, props in schema["properties"].items():
                    colname = props.get("csv_col", field)
                    record[field] = row_dict.get(colname, "")
                rows.append(record)
    else:
        logging.error(
            f"Unsupported file extension: {ext}. Only .csv and .xlsx are supported."
        )
        raise ValueError(f"Unsupported file extension: {ext}")
    return rows


def authenticate():
    """Authenticate against DataFed using env vars (DF_USER/DF_PASS) or interactive input (skipped in dry-run)."""
    if is_dry_run():
        logging.info("[dry-run] Skipping authentication")
        return None
    if API is None:
        raise RuntimeError(
            "datafed library not available; install `datafed` or use --dry-run"
        )
    username = os.getenv("DF_USER") or input("DataFed username upload: ").strip()
    password = os.getenv("DF_PASS") or getpass.getpass("DataFed password: ").strip()
    api = API()
    try:
        api.loginByPassword(username, password)
        logging.info("Successfully authenticated with DataFed user '%s'", username)
        return api
    except Exception as e:  # pragma: no cover
        logging.error(f"Authentication failed: {e}")
        raise RuntimeError(f"Authentication failed: {e}")


def update_record(df_api, rec, title, row_index):
    """
    Updates the metadata of an existing DataFed record identified by 'title'.
    Merges additional fields from the record (rec) into the existing metadata, skipping duplicates.
    row_index is used for logging (the row number in the CSV or text input).
    """
    try:
        # Attempt to find existing record in DataFed
        record_check = df_api.queryDirect(id=title)
        if not record_check or not record_check[0].item:
            logging.warning(
                f"Row {row_index}: Record '{title}' doesn't exist for updating. Skipping."
            )
            return
        existing_id = record_check[0].item[0].id
        logging.info(
            f"Row {row_index}: Found existing record ID '{existing_id}' for updating."
        )

        # Retrieve existing metadata
        existing_data = df_api.dataView(existing_id)
        current_metadata = {}
        if existing_data and existing_data[0].data[0].metadata:
            current_metadata_str = existing_data[0].data[0].metadata
            try:
                current_metadata = json.loads(current_metadata_str)
            except json.JSONDecodeError:
                pass

        # Merge in new fields from rec (skip duplicates)
        for k, v in rec.items():
            if k in current_metadata:
                logging.debug(
                    f"Row {row_index}: Duplicate field '{k}' found, skipping."
                )
            else:
                current_metadata[k] = v

        # Convert updated metadata to JSON string
        updated_meta_str = json.dumps(current_metadata)

        # Update record with new metadata
        df_api.dataUpdate(data_id=existing_id, metadata=updated_meta_str)
        logging.info(
            f"Row {row_index}: Updated record '{title}' (ID={existing_id}) with new metadata."
        )

    except Exception as e:
        logging.error(
            f"Row {row_index}: Failed to update existing record '{title}': {e}"
        )


def get_missing_files(
    df_api, rec_id: str, candidate_files: list[str], context: str | None = None
) -> list[str]:
    """
    Return a sub-list of candidate_files that are NOT already attached
    to the DataFed record rec_id (case-sensitive filename match).
    """
    try:
        reply = df_api.dataView(rec_id, context=context)
        existing = set()
        # print(f"reply: {reply[0]}")
        if reply and reply[0].data and reply[0].data[0].source:
            # Each file object usually has a .path field (full DataFed path)
            existing = {os.path.basename(f.path) for f in reply[0].data[0].source}
    except Exception as e:
        logging.warning(f"Could not inspect files for record {rec_id}: {e}")
        existing = set()

    return [f for f in candidate_files if os.path.basename(f) not in existing]


def process_records(
    df_api, records, schema, parent_collection, context, create_only=False
):
    """
    For each record defined in the CSV (interpreted via schema),
    create a DataFed record with metadata and dependencies.
    """
    # If schema has update_mode == True, we do an update scenario instead of creation
    update_mode = schema.get("update_record", False)
    record_tags = schema.get("tags", None)
    for i, rec in enumerate(records, start=2):
        # try:
        # Determine title from schema (e.g., using "output_file")
        title = rec.get(schema["title_from"], "")
        # Remove spaces from titles
        title = title.replace(" ", "-")

        print(f"Checking {title}.")
        # Handle bad titles (WIP)
        # TODO: Discuss handling these, seems like bad practice to include records with missing or re-used names
        skip_list = ["BLANK"]
        if title in skip_list or title == "":
            print("Missing title, moving to next record")
            continue  # Skip to the next record for now

        # If update mode is true, try to update existing records
        if update_mode:  # GOTCHA: update mode skips creation entirely; no file uploads attempted here.
            update_record(df_api, rec, title, i)
            continue  # skip the creation path

        # Handle record already existing in DataFed (skip)
        # TODO: Should existing records be skipped or updated?
        record_check = df_api.queryDirect(id=title, owner=context)
        if record_check and record_check[0].item:
            rec_id_exist = record_check[0].item[0].id
            logging.info(f"Record {title} already exists (ID {rec_id_exist}).")

            # Build list of files we expect for this record
            files_to_upload = []
            output_file = rec.get("output_file", "")

            if (
                not output_file
                and schema.get("schema_title") == "DNA Extractions Upload Schema"
            ):
                base = title.replace(".fastq.gz", "")
                files_to_upload = [f"{base}_R1_001.fastq.gz", f"{base}_R2_001.fastq.gz"]
            elif output_file:
                files_to_upload = [output_file]

            # Determine which of those files are NOT yet in DataFed
            missing = get_missing_files(
                df_api, rec_id_exist, files_to_upload, context=context
            )  # TODO: Deduplicate case-insensitively if DataFed treats filenames case-insensitively.

            # Upload only the missing ones
            for fname in missing:
                # local_path = os.path.join(os.getcwd(), "data/inputs/upload/", fname)
                # Build local path for file upload from configured / CLI provided reads directory
                local_path = os.path.join(
                    reads_dir, fname
                )  # GOTCHA: relies on global reads_dir set in run(); fragile if calling programmatically.
                if not os.path.exists(local_path):
                    logging.error(
                        f"Row {i}: Expected file '{fname}' not found at {local_path}."
                    )
                    continue
                try:
                    df_api.dataPut(rec_id_exist, local_path)
                    logging.info(
                        f"Row {i}: Uploaded missing file '{fname}' to record {rec_id_exist}."
                    )
                except Exception as e:
                    logging.error(f"Row {i}: Failed to upload '{fname}': {e}")
            # Skip record creation
            continue  # next CSV row
        else:
            print(f"Record for {title} doesn't exist, creating...")

        # title = transform_title(raw_title, schema.get("title_transform"))
        # If an alias isn't specified, set from record title
        alias = rec.get(
            "alias", title
        )  # GOTCHA: alias collisions not checked; server may reject duplicates.
        # Make sure the alias is more likely to be valid by removing spaces
        alias = alias.replace(" ", "-")
        meta_data = rec  # Use the entire record as metadata

        # Build dependencies from the record_relationship_field and
        # relationship_type
        deps = []
        rel_field = schema.get("record_relationship_field", None)
        if rel_field:
            relationship_type = schema.get("relationship_type")
            # Handle multiple relationships
            if rel_field in rec and isinstance(rec[rel_field], list):
                for rel in rec[rel_field]:
                    # Query DataFed for the associated record
                    rel_query = df_api.queryDirect(id=rel, owner=context)
                    if rel_query and len(rel_query) > 0 and rel_query[0].item:
                        rel_id = rel_query[0].item[0].id
                        deps.append([relationship_type, rel_id])
                    else:
                        logging.warning(
                            f"Relationship '{rel}' not found in DataFed. Skipping this dependency."
                        )
            else:  # Single relationship
                rel_value = rec[rel_field]
                rel_query = df_api.queryDirect(id=rel_value, owner=context)
                if rel_query and len(rel_query) > 0 and rel_query[0].item:
                    rel_id = rel_query[0].item[0].id
                    deps.append([relationship_type, rel_id])
                else:
                    logging.warning(
                        f"Relationship '{rel_value}' not found in DataFed. Skipping this dependency."
                    )

        # Create the DataFed record call
        if (
            deps
        ):  # GOTCHA: dependency resolution assumes referenced records already exist.
            record = df_api.dataCreate(
                title=title,
                alias=alias,
                metadata=json.dumps(meta_data),
                tags=record_tags,
                parent_id=parent_collection,
                deps=deps,
            )
        else:  # No dependencies or none found
            record = df_api.dataCreate(
                title=title,
                alias=alias,
                metadata=json.dumps(meta_data),
                tags=record_tags,
                parent_id=parent_collection,
            )
        if not record or not record[0].data:
            logging.error(f"Row {i}: Failed to create record for title '{title}'.")
            continue
        record_id = record[0].data[0].id
        logging.info(f"Row {i}: Created record '{title}' with ID {record_id}.")

        if create_only:
            continue

        # Upload the output file
        output_file = rec.get("output_file", "")
        # If no output_file is specified and we're working with the dna extraction schema,
        # we guess the actual file names for R1 and R2 based on the record title.
        if (
            not output_file
            and schema.get("schema_title") == "DNA Extractions Upload Schema"
        ):  # GOTCHA: Assumes naming convention: <title>_R1/_R2_001.fastq.gz derived from title sans .fastq.gz.
            # Remove ".fastq.gz" to get a base
            base_name = title.replace(".fastq.gz", "").strip()
            # Build the R1 and R2 filenames
            read_files = [
                f"{base_name}_R1_001.fastq.gz",
                f"{base_name}_R2_001.fastq.gz",
            ]
            # Attempt to upload both
            for rf in read_files:
                # upload_path = os.path.join(os.getcwd(), "data/upload/", rf)
                # Build upload path (reads directory + resolved output_file)
                upload_path = os.path.join(
                    reads_dir, output_file
                )  # TODO: Likely bug: should use rf not output_file when iterating inferred read_files.

                if not os.path.exists(upload_path):
                    logging.error(
                        f"Row {i}: Output file '{rf}' does not exist at '{upload_path}'. Skipping."
                    )
                    continue
                try:
                    df_api.dataPut(record_id, upload_path)
                    logging.info(
                        f"Row {i}: Uploaded file '{rf}' to record ID {record_id}."
                    )
                except Exception as e:
                    logging.error(f"Row {i}: Failed to upload file '{rf}': {e}")
        else:
            # If output_file is defined or we're not using the "Sequence Metadata Upload" schema,
            # proceed with the original single-file approach.
            # upload_path = os.path.join(os.getcwd(), "data/upload/", output_file)
            upload_path = os.path.join(reads_dir, output_file)

            if not os.path.exists(upload_path):
                logging.error(
                    f"Row {i}: Output file '{output_file}' does not exist at '{upload_path}'."
                )
                continue
            df_api.dataPut(record_id, upload_path)
            logging.info(
                f"Row {i}: Uploaded file '{output_file}' to record ID {record_id}."
            )


def is_dry_run():
    """Return True when dry-run mode active via env var or CLI flag-set attribute."""
    return os.environ.get(DRY_RUN_ENV) == "1" or getattr(is_dry_run, "_flag", False)


def run(argv=None):
    """CLI entry: parse args, load config+schema, optionally perform uploads (respecting dry-run)."""
    parser = argparse.ArgumentParser(
        description="Upload data and metadata to DataFed using a JSON schema."
    )
    parser.add_argument(
        "--config",
        default="config/DataFedSettings.cfg",
        help="Path to the configuration file.",
    )
    parser.add_argument("--schema", required=True, help="Path to the JSON schema file.")
    parser.add_argument(
        "--create-only",
        action=argparse.BooleanOptionalAction,
        help="Only create the DataFed records without uploading files.",
    )
    parser.add_argument(
        "--reads-dir",
        default="Data/reads",
        help="Directory containing read / data files (defaults to Data/reads).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Load schema & config only; skip authentication & uploads",
    )
    args = parser.parse_args(argv)

    if args.dry_run:
        is_dry_run._flag = True  # type: ignore

    global reads_dir  # GOTCHA: Avoid global if refactoring into library usage; pass as parameter instead.
    reads_dir = args.reads_dir
    if not os.path.exists(reads_dir):
        logging.warning(
            f"Reads directory '{reads_dir}' does not exist (proceeding; uploads may fail)"
        )

    if not os.path.exists(args.config):
        print(f"No config file found at '{args.config}'.")
        create_config(args.config)
        print("Please update the configuration file and rerun the script.")
        return 2
    if load_workbook is None:
        logging.warning("openpyxl not available; XLSX parsing will fail if needed")
    config_settings = read_config(args.config)
    log_file = os.path.join("data", "logs", "upload.log")
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    setup_logging(log_file)
    schema = load_schema(args.schema)
    batch_file = schema["batch_file"]
    if not os.path.exists(batch_file):
        logging.error(f"Batch upload file '{batch_file}' not found.")
        return 3
    read_row_start = schema.get("read_row_start", 1)
    records = parse_data_with_schema(batch_file, schema, read_row_start)
    if is_dry_run():  # GOTCHA: Parsed records retained only in memory; no optional export for inspection.
        logging.info("[dry-run] Parsed schema & batch file (%d records)", len(records))
        return 0
    df_api = authenticate()
    globus_uuid = config_settings.get(CONFIG_SECTION, "globus_uuid")
    df_api.endpointDefaultSet(globus_uuid)
    logging.info(f"Set default Globus endpoint to UUID: {globus_uuid}")
    context = config_settings.get(CONFIG_SECTION, "context")
    df_api.setContext(context)
    logging.info(f"Set DataFed context to: {context}")
    parent_collection = config_settings.get(CONFIG_SECTION, "parent_collection")
    logging.info("Setting collection to: %s", parent_collection)
    process_records(
        df_api,
        records,
        schema,
        parent_collection,
        context,
        create_only=args.create_only,
    )
    logging.info("DataFed upload workflow completed successfully.")
    return 0


def main(argv=None, embedded=False):
    """Module main wrapper: raises on non-zero exit when embedded (programmatic use)."""
    code = run(argv)
    if embedded or os.environ.get("ECOKMER_EMBEDDED") == "1":
        if code != 0:
            raise RuntimeError(f"datafed_csv_upload failed (code={code})")
    else:
        import sys

        sys.exit(code)


if __name__ == "__main__":  # pragma: no cover
    main()
